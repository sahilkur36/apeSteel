"""Faithfully dump an engineer Excel workbook to the golden JSON schema.

Produces the same structure used by the Chapter-E external anchor
(``tests/golden/data/_compression_excel_raw.json``):

    {
      "source_file": "<basename>",
      "defined_names": {<name>: <attr_text>, ...},
      "sheets": {
        "<sheet>": {
          "max_row": int, "max_col": int,
          "cells": [{"coord","row","col","formula","value"}, ...]
        }, ...
      }
    }

The workbook is opened twice: ``data_only=False`` for the *formula*
text and ``data_only=True`` for Excel's last-saved *cached value*.
Only non-empty cells are emitted.  Sheets are allow-listed on the
command line so huge embedded lookup tables (e.g. an AISC shapes
database) are excluded from the fixture.

This is a *provenance* tool: the JSON it writes is the external anchor
the golden test bit-matches, so a teammate can re-dump and diff.

Usage
-----
    python tools/dump_excel_workbook.py <workbook> <out.json> SHEET [SHEET ...]
"""

from __future__ import annotations

import json
import sys
import warnings
from pathlib import Path
from typing import Any

import openpyxl


def _coerce(obj: Any) -> Any:
    """JSON-safe scalar.  openpyxl yields ArrayFormula / DataTableFormula
    objects for spilled / data-table cells; serialise those (and dates)
    as their string form so the fixture stays a plain JSON document."""
    if obj is None or isinstance(obj, (str, int, float, bool)):
        return obj
    return str(obj)


def dump_workbook(src: Path, sheets: list[str]) -> dict[str, Any]:
    """Return the golden-schema dict for ``src`` restricted to ``sheets``."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb_f = openpyxl.load_workbook(src, data_only=False, keep_vba=True)
        wb_v = openpyxl.load_workbook(src, data_only=True, keep_vba=True)

    defined_names: dict[str, str] = {}
    for name, defn in wb_f.defined_names.items():
        defined_names[name] = str(defn.value)

    out_sheets: dict[str, Any] = {}
    for title in sheets:
        if title not in wb_f.sheetnames:
            raise SystemExit(f"sheet {title!r} not in workbook {wb_f.sheetnames}")
        ws_f = wb_f[title]
        ws_v = wb_v[title]
        cells: list[dict[str, Any]] = []
        for row in ws_f.iter_rows():
            for cell in row:
                formula = _coerce(cell.value)
                value = _coerce(ws_v[cell.coordinate].value)
                if formula is None and value is None:
                    continue
                cells.append(
                    {
                        "coord": cell.coordinate,
                        "row": cell.row,
                        "col": cell.column,
                        "formula": formula,
                        "value": value,
                    }
                )
        out_sheets[title] = {
            "max_row": ws_f.max_row,
            "max_col": ws_f.max_column,
            "cells": cells,
        }

    return {
        "source_file": src.name,
        "defined_names": defined_names,
        "sheets": out_sheets,
    }


def main(argv: list[str]) -> int:
    if len(argv) < 4:
        print(__doc__)
        return 2
    src = Path(argv[1])
    out = Path(argv[2])
    sheets = argv[3:]
    data = dump_workbook(src, sheets)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    n = sum(len(s["cells"]) for s in data["sheets"].values())
    print(f"wrote {out} : {len(data['sheets'])} sheet(s), {n} non-empty cells")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
